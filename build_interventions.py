#!/usr/bin/env python3
"""One-off processor: turns PACIDA's real project-summary spreadsheet + @Glance
presentation into assets/interventions.json for the site. Source Office files stay
outside the repo (Downloads/UNICEF) — only this derived, public-appropriate JSON
(titles, thematic areas, donors, years, locations, published achievement stats)
gets committed. No per-project grant amounts are published (those read as internal
financial detail, not something in the externally-shared @Glance deck).
"""
import json, re, os, datetime, random

SITE = os.path.dirname(os.path.abspath(__file__))
SRC = r"C:\Users\l.muchemi\Downloads\UNICEF\PACIDA Project Summary.xlsx"

import openpyxl

# ---------- 1. pull every project row ----------
wb = openpyxl.load_workbook(SRC, data_only=True)
raw = []
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    header_idx, layout = None, None
    for i, row in enumerate(rows):
        if row and row[0] in ("Year of Commencement", "Year of commencement"):
            header_idx = i
            layout = "full" if (len(row) > 2 and row[2] == "Back Donor") else "noback"
            break
        if row and row[0] == "Year":
            header_idx, layout = i, "noback"
            break
    if header_idx is None:
        continue
    for row in rows[header_idx + 1:]:
        if not row or not any(row):
            continue
        if layout == "full":
            year, donor, back, thematic, title, dur, start, end = (list(row) + [None] * 8)[:8]
        else:
            year, donor, thematic, title, dur, start, end = (list(row) + [None] * 7)[:7]
        if not title:
            continue
        raw.append(dict(sheet=name.strip(), year=year, donor=donor, thematic=thematic,
                         title=str(title).strip(), duration=dur, start=start, end=end))

# ---------- 2. normalize thematic category ----------
THEME_MAP = [
    (r"emergency|humanitarian", "Emergency", "#ED1C24"),
    (r"anticipatory", "Disaster Risk Reduction", "#E8834A"),
    (r"disaster risk|\bdrr\b", "Disaster Risk Reduction", "#E8834A"),
    (r"wash", "WASH", "#6FA3B4"),
    (r"nutrition|health", "Health & Nutrition", "#D66FB0"),
    (r"education", "Education", "#F0B22E"),
    (r"climate", "Climate Change", "#5FBF8F"),
    (r"peace|governance|conflict", "Peace & Governance", "#9C87D9"),
    (r"livelihood|livestock", "Livelihood", "#34B44B"),
]
def normalize_theme(raw_theme):
    t = (raw_theme or "").lower()
    for pattern, label, color in THEME_MAP:
        if re.search(pattern, t):
            return label, color
    return "Other", "#9A8F70"

# ---------- 3. status from end date ----------
NOW = datetime.date(2026, 7, 29)
def to_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None

# ---------- 4. gazetteer for location tagging (PACIDA's 4 real operational areas) ----------
GAZETTEER = {
    "marsabit": {
        "sites": ["Marsabit town", "Moyale", "North Horr", "Kalacha", "Maikona", "Loiyangalani",
                  "Laisamis", "Korr", "Illeret", "Sololo", "Turbi", "Hurri Hills", "Kargi", "Dukana"],
        "county_names": ["Marsabit"],
    },
    "samburu": {
        "sites": ["Maralal", "Baragoi", "Wamba", "Archer's Post", "South Horr", "Suguta Marmar", "Sereolipi"],
        "county_names": ["Samburu"],
    },
    "isiolo": {
        "sites": ["Isiolo town", "Isiolo", "Merti", "Garbatulla", "Kinna", "Oldonyiro", "Sericho", "Ngare Mara"],
        "county_names": ["Isiolo"],
    },
    "borena": {
        "sites": ["Yabelo", "Moyale (ET)", "Mega", "Dubluk", "Teltele", "Dillo", "Dilo", "Arero", "Hidi Lola",
                  "Turmi", "Omorate", "South Omo"],
        "county_names": ["Ethiopia", "South Ethiopia", "Southern Ethiopia", "Somalia"],
    },
}
# longest names first so e.g. "Moyale (ET)" matches before bare "Moyale"
ALL_NAMES = []
for slug, g in GAZETTEER.items():
    for n in g["sites"]:
        ALL_NAMES.append((n, slug, "site"))
    for n in g["county_names"]:
        ALL_NAMES.append((n, slug, "county"))
ALL_NAMES.sort(key=lambda x: -len(x[0]))

def find_locations(title):
    hits = []  # (slug, kind, name)
    remaining = title
    for name, slug, kind in ALL_NAMES:
        if re.search(r'\b' + re.escape(name) + r'\b', remaining, re.IGNORECASE):
            hits.append((slug, kind, name))
            remaining = re.sub(re.escape(name), "", remaining, flags=re.IGNORECASE)
    # dedupe per slug, prefer site-level over county-level
    by_slug = {}
    for slug, kind, name in hits:
        if slug not in by_slug or (by_slug[slug][0] == "county" and kind == "site"):
            by_slug[slug] = (kind, name)
    return by_slug  # {slug: (kind, name)}

counties = json.load(open(os.path.join(SITE, "counties.json"), encoding="utf-8"))
boundaries = json.load(open(os.path.join(SITE, "assets", "boundaries.json"), encoding="utf-8"))
SITE_COORDS = {}
for slug, r in counties.items():
    for s in r["sites"]:
        SITE_COORDS[(slug, s[0])] = (s[1], s[2])
COUNTY_HQ = {slug: (r["hq"]["lat"], r["hq"]["lon"]) for slug, r in counties.items()}

# a few real places named in project titles/offices that aren't in the curated "sites" lists
MANUAL_COORDS = {
    "Hurri Hills": (2.75, 37.75), "Kargi": (2.50, 37.57), "Dukana": (4.16, 37.11),
    "Turmi": (4.95, 36.50), "Omorate": (4.83, 36.05), "Dilo": (4.20, 37.73),
}


def point_in_ring(lon, lat, ring):
    inside = False
    n = len(ring)
    for i in range(n):
        x1, y1 = ring[i]
        x2, y2 = ring[(i + 1) % n]
        if ((y1 > lat) != (y2 > lat)) and (lon < (x2 - x1) * (lat - y1) / (y2 - y1 + 1e-12) + x1):
            inside = not inside
    return inside


def point_in_geom(lon, lat, geom):
    if geom["type"] == "Polygon":
        return point_in_ring(lon, lat, geom["coordinates"][0])
    else:  # MultiPolygon
        return any(point_in_ring(lon, lat, poly[0]) for poly in geom["coordinates"])


def county_bbox(slug):
    geom = boundaries.get(slug)
    if not geom:
        return None
    pts = []

    def walk(c):
        if isinstance(c[0], (int, float)):
            pts.append(c)
        else:
            for x in c:
                walk(x)
    walk(geom["coordinates"])
    lons = [p[0] for p in pts]
    lats = [p[1] for p in pts]
    return min(lons), max(lons), min(lats), max(lats)


_BBOX_CACHE = {}

def scattered_county_point(slug, seed_key):
    """A real, plausible point somewhere inside the county's actual boundary — used for
    projects whose title only names the county, not a specific site, so they don't all
    stack on the exact same HQ pixel on the map."""
    geom = boundaries.get(slug)
    if not geom:
        return COUNTY_HQ.get(slug)
    if slug not in _BBOX_CACHE:
        _BBOX_CACHE[slug] = county_bbox(slug)
    bbox = _BBOX_CACHE[slug]
    if not bbox:
        return COUNTY_HQ.get(slug)
    minlon, maxlon, minlat, maxlat = bbox
    rnd = random.Random(seed_key)
    for _ in range(40):
        lon = rnd.uniform(minlon, maxlon)
        lat = rnd.uniform(minlat, maxlat)
        if point_in_geom(lon, lat, geom):
            return (lat, lon)
    return COUNTY_HQ.get(slug)


# projects whose title names no specific place at all (regional / cross-cutting /
# multi-county programmes) still get mapped — scattered across PACIDA's real
# operational area, weighted by each area's households (a proxy for scale of
# reach), and tagged kind="regional" so the frontend can render/label them
# honestly as approximate rather than a precise project site.
PACIDA_AREA_SLUGS = ["marsabit", "samburu", "isiolo", "borena"]
_area_weights = [counties[s]["households"] for s in PACIDA_AREA_SLUGS]
_area_total = sum(_area_weights)
_area_cum = []
_c = 0
for _w in _area_weights:
    _c += _w
    _area_cum.append(_c / _area_total)


def pick_regional_slug(seed_key):
    rnd = random.Random(seed_key)
    r = rnd.random()
    for slug, cum in zip(PACIDA_AREA_SLUGS, _area_cum):
        if r <= cum:
            return slug
    return PACIDA_AREA_SLUGS[-1]


def coords_for(slug, kind, name, seed_key=None):
    if kind == "site":
        if (slug, name) in SITE_COORDS:
            return SITE_COORDS[(slug, name)]
        if name in MANUAL_COORDS:
            return MANUAL_COORDS[name]
        # site name variants that don't exactly match the sites list (e.g. "Dilo" vs "Dillo", "Isiolo" vs "Isiolo town")
        for (s2, n2), (lat, lon) in SITE_COORDS.items():
            if s2 == slug and n2.lower().startswith(name.lower()[:5]):
                return (lat, lon)
    if kind == "county" and seed_key is not None:
        return scattered_county_point(slug, seed_key)
    return COUNTY_HQ.get(slug)

projects = []
loc_hits, no_loc = 0, 0
for i, p in enumerate(raw):
    theme_label, theme_color = normalize_theme(p["thematic"])
    end_d = to_date(p["end"])
    status = "ongoing" if (end_d is None or end_d >= NOW) else "completed"
    locs = find_locations(p["title"])
    loc_list = []
    if locs:
        loc_hits += 1
        for s, (k, n) in locs.items():
            c = coords_for(s, k, n, seed_key=f"proj-{i}-{s}")
            if c:
                loc_list.append({"slug": s, "kind": k, "name": n, "lat": c[0], "lon": c[1]})
    else:
        no_loc += 1
        rslug = pick_regional_slug(f"regional-{i}")
        c = scattered_county_point(rslug, f"regional-{i}")
        if c:
            loc_list.append({"slug": rslug, "kind": "regional", "name": "Regional programme", "lat": c[0], "lon": c[1]})
    projects.append(dict(
        id=f"proj-{i}", year=p["year"], donor=(p["donor"] or "").strip(),
        theme=theme_label, theme_color=theme_color, title=p["title"],
        duration=(p["duration"] or "").strip() if isinstance(p["duration"], str) else p["duration"],
        start=str(p["start"])[:10] if p["start"] else None,
        end=str(p["end"])[:10] if p["end"] else None,
        status=status,
        locations=loc_list,
    ))

print(f"projects: {len(projects)}  precisely-sited: {loc_hits}  scattered-regional (title names no specific place): {no_loc}")

# ---------- 5. curated achievement stats (from PACIDA's own externally-shared @Glance deck) ----------
ACHIEVEMENTS = [
    {"n": "123,902", "label": "people gained access to safe water"},
    {"n": "26,034", "label": "households on nutrition-sensitive cash transfers (~156,200 people)"},
    {"n": "640,692", "label": "livestock vaccinated, treated or dewormed"},
    {"n": "50,447", "label": "households reached with One Health crisis support"},
    {"n": "10,138", "label": "households received nutritious food baskets"},
    {"n": "27,913", "label": "students supported through school feeding (12,206 boys / 11,647 girls)"},
    {"n": "20", "label": "boreholes drilled + 11 rehabilitated (52,298 people benefiting)"},
    {"n": "30+", "label": "water sources solarized; 25 shallow wells improved; 5 water supply systems built"},
    {"n": "17,100", "label": "households supported with livestock feeds"},
    {"n": "5,200+", "label": "community health volunteers & pregnant/lactating women trained (SBCC)"},
    {"n": "2,940", "label": "shoats + 40 camels distributed to vulnerable women's groups"},
    {"n": "2", "label": "cross-border peace accords supported (Daasanach–Gabra; Dilo–Dukana review)"},
    {"n": "1,935", "label": "teenage girls reached with menstrual hygiene management support"},
    {"n": "128", "label": "water-trucking trips, delivering emergency water to 71,604 people"},
]
CHALLENGES = [
    {"challenge": "Recurring droughts and climate shocks continue to undermine livelihoods",
     "mitigation": "Scale up Anticipatory Action, climate-smart livelihoods, and resilient natural resource management."},
    {"challenge": "Food insecurity and acute malnutrition remain critical in ASAL areas",
     "mitigation": "Strengthen integrated health, nutrition, food security, and cash assistance interventions."},
    {"challenge": "Resource-based conflicts and insecurity persist across pastoralist zones",
     "mitigation": "Promote community-led peacebuilding, conflict resolution, and cross-border collaboration."},
    {"challenge": "High illiteracy and low education access and retention among nomadic communities",
     "mitigation": "Expand access to inclusive, quality, and resilient education for pastoralist children."},
    {"challenge": "Water scarcity and weak WASH infrastructure continue to affect communities",
     "mitigation": "Invest in climate-resilient water systems and sustainable WASH services."},
    {"challenge": "High donor dependency remains a major institutional risk",
     "mitigation": "Diversify funding sources while strengthening institutional sustainability and partnerships."},
    {"challenge": "Drug and substance misuse, especially among youth",
     "mitigation": "Empower youth through prevention, life skills, livelihoods, and community support programmes."},
    {"challenge": "Political dynamics and governance challenges",
     "mitigation": "Strengthen stakeholder engagement, adaptive programming, and institutional risk management."},
]
# ---------- 5b. real target-population / beneficiary figures ----------
# Sourced from PACIDA's own narrative/progress/final reports (not the financial
# project-summary register, which has no beneficiary field). One clean summary
# per project, condensed from the full report text — verified against the
# register by donor/title/year. Projects with only fragmented activity-level
# counts (no reliable cumulative total) are deliberately left out rather than
# forcing a number.
POPULATIONS = {
    "proj-0": "231,240 people reached (136,210 in Marsabit) with livestock, cash & WASH support — consortium led by Concern Worldwide",
    "proj-1": "1,294 people reached (2023–24) through women's climate-resilience groups, training and gender/peace forums, Turbi Ward",
    "proj-2": "8 community organisations awarded UNDP-GEF small grants (~USD180,000); 800 students reached via school forest sensitization",
    "proj-3": "400 households reached with unconditional cash assistance (148 male, 252 female-headed persons; 112 with disabilities)",
    "proj-4": "5,100 people reached (1,400 male, 3,700 female) with food aid, school feeding and water trucking, Moyale/Dire Woredas",
    "proj-12": "12,000 people targeted via rehabilitated boreholes; 9,600 reached through One Health sensitization, plus indirect reach across Gabbra & Borena communities",
    "proj-13": "8,220 households targeted (+41,100 indirect); at endline 4,672 households achieved a 25% income increase, Kajiado/Marsabit/Turkana",
    "proj-14": "392 women & youth trained on county governance processes; 58 WDPCs/WRUAs/CSOs engaged in fiscal-planning review, Samburu",
    "proj-15": "243 participants reached in one quarter across water-resource, climate-smart agriculture and WASH forums, Samburu",
    "proj-16": "480 women completed business-skills training; over 85% of surveyed households reported increased income, 4 sub-counties",
    "proj-17": "2,926 households reached via cross-border livestock vaccination campaigns (215,834 livestock treated); 260 VSLA participants, 80% women",
    "proj-18": "21 CSOs engaged across 13 counties; 1,300+ people trained on rights-based approaches",
    "proj-22": "3,310 households reached with cash transfers — 2,580 in Marsabit & Samburu, 730 in Borena Zone",
    "proj-23": "150 households restocked with goats; 300 households received cash transfers after the 2024 floods, Moyale",
    "proj-24": "4,704 people reached with safe drinking-water access and hygiene training across Turkana & Kajiado",
    "proj-26": "7,670 households (41,017 people) targeted for WASH support across 4 counties; 200 households restocked",
    "proj-27": "Cross-border livestock-market and youth-TVET forums reaching 30–90 participants per session, Forolle–Dirre corridor",
    "proj-33": "800 households (4,800 people) targeted for six cycles of multi-purpose cash transfers",
    "proj-36": "90 community members trained on climate-adaptation DRR across 6 sites in Marsabit & Ethiopia",
    "proj-38": "2,524 people trained in climate-resilient food production (of a 3,000 target); 1,099 households received cash transfers",
    "proj-49": "400 flood-affected households reached with cash assistance via M-Pesa, Kargi & Maikona",
    "proj-51": "400 pupils supported at Tiigo Primary School; 37 high schools received Form 1 fee support",
    "proj-55": "200 persons with disabilities each received livestock restocking; 200 farmers trained, Hurri Hills/Maikona",
    "proj-60": "700 households supported with 3 months of food supplies, Maikona & Dukana wards",
    "proj-61": "283 households received cash transfers; 220 households (1,320 people) reached with water trucking, North Horr",
    "proj-63": "418 households (2,508 people) targeted for drought assistance, Laisamis",
    "proj-64": "709 households (3,815 people) reached with cash and hygiene support, North Horr (609 main + 100-household top-up)",
    "proj-68": "1,796 households on the cash-transfer caseload, Kenya COVID-19 Fund response",
    "proj-71": "23,000 people targeted for cash and water-trucking assistance across 5 wards, Marsabit",
    "proj-72": "390 women reached with SRHR information; 125 accessed services; 60 youth trained as climate-justice champions, Samburu",
    "proj-80": "1,060 people reached with early cash assistance, Maikona & Kargi",
    "proj-81": "750 people trained in Good Agricultural Practices; 300 households received cash transfers; 382 shoats distributed",
    "proj-85": "425 pupils supported at Tiigo School (up from 365); new dormitory adds capacity for 80 more",
    "proj-93": "30,000 people served by 16 community-managed boreholes; oxygen support to 10 health facilities during COVID-19",
    "proj-94": "340 people reached through WCCPC formation forums and public-participation sessions on Ward Development Plans",
    "proj-101": "503 people reached through ward-based dissemination of Marsabit's County Climate Change Policy",
    "proj-116": "70 community volunteers trained in disease surveillance across 4 sub-counties, with Oxfam & ALDEF",
    "proj-125": "Institutional capacity-building for 40 PACIDA staff and 7 Board members",
    "proj-128": "3,000 pastoralists (500 families) and 20,000 livestock targeted, Chaffa Chachane rangelands",
    "proj-134": "367 people reached through DRR/contingency-planning forums; rainwater-harvesting structure built, Horonder",
    "proj-137": "1,975 households (11,850 people) reached with food supplies; 1,089 households via water trucking",
    "proj-138": "Solar-powered borehole at Balesa supporting 3,200 households and ~15,000 livestock",
    "proj-140": "Over 1,900 learners and 600 parents benefited; 159 students received high-school fee bursaries",
    "proj-144": "4,019 households (24,114 people) reached with food, water and school meals, North Horr",
    "proj-146": "5,000 people and 60,000 livestock targeted, Turbi division",
    "proj-149": "24,472 people reached with food, school feeding and water-trucking support",
    "proj-150": "4,000 households (24,000 people) targeted for food distribution; 2,003 households via water trucking",
    "proj-164": "2,326 households reached with food support (above the 1,750 target); 7 shallow wells rehabilitated",
}

OFFICES = [
    {"name": "PACIDA HQ", "slug": "marsabit", "site": "Marsabit town", "note": "Head office, warehouse (owned)"},
    {"name": "Moyale field office", "slug": "marsabit", "site": "Moyale", "note": "Kenya field office"},
    {"name": "Hurri Hills field office", "slug": "marsabit", "site": "Hurri Hills", "note": "Currently not operational (funding constraints)"},
    {"name": "Illeret field office", "slug": "marsabit", "site": "Illeret", "note": "Currently not operational (funding constraints)"},
    {"name": "Turbi field office", "slug": "marsabit", "site": "Turbi", "note": "Currently not operational (funding constraints)"},
    {"name": "Samburu county office", "slug": "samburu", "site": "Maralal", "note": "County office"},
    {"name": "Isiolo county office", "slug": "isiolo", "site": "Isiolo town", "note": "County office"},
    {"name": "Moyale (Ethiopia) office", "slug": "borena", "site": "Moyale (ET)", "note": "Ethiopia country office"},
    {"name": "Turmi/Omorate office", "slug": "borena", "site": "Turmi", "note": "South Omo field presence"},
]

for o in OFFICES:
    c = coords_for(o["slug"], "site", o["site"])
    if c:
        o["lat"], o["lon"] = c

for proj in projects:
    pop = POPULATIONS.get(proj["id"])
    if pop:
        proj["population"] = pop

THEMES = sorted({(p["theme"], p["theme_color"]) for p in projects}, key=lambda x: x[0])

# ---------- 6. real donor/partner names — merged from two sources ----------
# (a) the project register's own "donor" field (complete, but full of typo/
#     casing variants from years of manual entry — "Caritus Austria",
#     "Concer Worldwide", "WHH" etc. all mean the same organisation), and
# (b) PACIDA's Partnership Documents folder names (organisations with a
#     formal partnership record but not necessarily a completed/dated grant
#     row yet). Only organisation names are used from either source — no
#     contract, budget, audit or compliance content.
DONOR_CANON = {
    "acdi/voca": "ACDI/VOCA", "acted": "ACTED", "adeso": "ADESO", "aldef": "ALDEF",
    "amref maanisha": "AMREF Maanisha", "accord": "Accord", "action aid": "ActionAid",
    "asociatia sharing love": "Sharing Love", "bild": "BILD",
    "caritas austria": "Caritas Austria", "caritas ausria": "Caritas Austria",
    "caritus austria": "Caritas Austria", "caritas austria-": "Caritas Austria",
    "caritas austria-pclp": "Caritas Austria", "caritas bolzano": "Caritas Bolzano",
    "cbm": "CBM", "cewer": "CEWER", "cordaid": "Cordaid", "cordaid/echo": "Cordaid",
    "chemonics kenya limited": "Chemonics Kenya", "chemonics kenya": "Chemonics Kenya",
    "chistian aid": "Christian Aid", "christian aid": "Christian Aid",
    "concer worldwide": "Concern Worldwide", "concern worldwide": "Concern Worldwide",
    "concern world wide": "Concern Worldwide", "concernworldwide": "Concern Worldwide",
    "dka austria": "DKA Austria", "dka": "DKA Austria", "dorcas": "Dorcas", "drc": "DRC",
    "fao-101": "FAO", "fao-105": "FAO", "un-fao": "FAO", "unfao": "FAO",
    "un - fao (food and agriculture organizationof the united nations)": "FAO",
    "fhi": "FHI 360", "helvetas": "Helvetas", "hivos": "HIVOS",
    "horizon 3000": "Horizont3000", "horizont 3000": "Horizont3000",
    "instiglio": "Instiglio", "ipas africa alliance": "Ipas Africa Alliance", "ipas": "Ipas Africa Alliance",
    "jica": "JICA", "kcdf": "KCDF", "kindermissionswerk": "Kindermissionswerk", "knh": "Kindernothilfe (KNH)",
    "kenya covid - 19 fund (bega kwa bega)": "Bega Kwa Bega (Kenya COVID-19 Fund)",
    "mercycorps": "Mercy Corps", "mercy corps": "Mercy Corps",
    "miral foundation": "MIRAL Foundation", "miral": "MIRAL Foundation",
    "malteser international": "Malteser International", "miserior": "Misereor", "misereor": "Misereor",
    "mwa": "MWA-STAWI", "mwa-stawi": "MWA-STAWI", "near": "NEAR", "oxfam": "Oxfam",
    "ofda": "USAID/OFDA", "pelum": "PELUM", "save the children": "Save the Children",
    "snv": "SNV", "stp": "STP", "towa": "TOWA", "undp": "UNDP", "unops": "UNOPS",
    "usaid": "USAID", "uwezo-kenya": "Uwezo Kenya", "uwezo": "Uwezo Kenya",
    "waterfund": "WaterFund",
    "whh": "Welthungerhilfe", "welthunger hilfe": "Welthungerhilfe",
    "welthunger hilfe(bmz)": "Welthungerhilfe", "welthunger hilfe(giz)": "Welthungerhilfe",
    "welthungerhilfe": "Welthungerhilfe",
    "act alliance": "ACT Alliance", "care international": "CARE International",
    "cfp (condivisione fra 1 propoli)": "CFP (Condivisione fra i Popoli)",
    "caritas germany": "Caritas Germany", "caritas german": "Caritas Germany",
    "caritas german(bmz)": "Caritas Germany", "caritas germany (bmz-s)": "Caritas Germany",
}
PARTNER_FOLDERS = [  # Partnership Documents folder names not already covered above
    "ACT Alliance", "CARE International", "CFP (Condivisione fra i Popoli)", "DRC",
    "Helvetas", "Instiglio", "Kindernothilfe (KNH)", "PELUM", "Save the Children",
    "STP", "USAID",
]
partner_set = set()
for p in projects:
    raw_name = (p["donor"] or "").strip()
    if not raw_name:
        continue
    canon = DONOR_CANON.get(raw_name.lower())
    partner_set.add(canon if canon else raw_name)
for name in PARTNER_FOLDERS:
    partner_set.add(name)
PARTNERS = sorted(partner_set, key=str.lower)

out = dict(
    generated=str(NOW), total_projects=len(projects), years_active=NOW.year - 2010,
    projects=projects, achievements=ACHIEVEMENTS, challenges=CHALLENGES, offices=OFFICES,
    themes=[{"label": t, "color": c} for t, c in THEMES], partners=PARTNERS,
)
open(os.path.join(SITE, "assets", "interventions.json"), "w", encoding="utf-8").write(
    json.dumps(out, ensure_ascii=False, indent=1)
)
print("wrote assets/interventions.json")
